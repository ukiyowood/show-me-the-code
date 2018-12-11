import openpyxl
import json

def txtParser(txt):
    with open(txt,'r',encoding='utf-8') as f:
        dict = json.load(f)
    return dict

def excelHandler(dict,excel):
    try:
        wb = openpyxl.load_workbook(excel)
    except IOError as e:
        print(e)
        wb = openpyxl.Workbook()
        wb.save(excel)
        wb = openpyxl.load_workbook(excel)
    else:
        sheet = wb.active
        sheet.title = excel
        for key,word in dict.items():
            sheet.cell(row=int(key),column=1,value=key)
            for i in range(0,len(word)):
                sheet.cell(row=int(key),column=i+2,value=word[i])
        wb.save(filename=excel)
        print('ok')

if __name__ == "__main__":
    txt = 'student.txt'
    data = txtParser(txt)
    excelHandler(data,'student.xlsx')